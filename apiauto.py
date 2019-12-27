import openpyxl
import requests


def web_get(login_para, web_url):
    response = requests.get(web_url, params=login_para)
    return response.json()


def read_test(xlsx, selected_form):
    wa = openpyxl.load_workbook(xlsx)
    form = wa[selected_form]
    maxrow = form.max_row
    list_test = []
    for i in range(2, maxrow + 1):
        dict_test = dict(
            num=form.cell(row=i, column=1).value,
            url=form.cell(row=i, column=5).value,
            data=form.cell(row=i, column=6).value,
            expected_result=form.cell(row=i, column=7).value)
        list_test.append(dict_test)
    return list_test


def write_result(filename, sheetname, row, column, final_result):
    wb = openpyxl.load_workbook(filename)
    sheetname = wb[sheetname]
    sheetname.cell(row, column).value = final_result
    wb.save(filename)


def exec_cases(filename, formname):
    cases = read_test(filename, formname)
    success = 0
    failed = 0
    for case in cases:
        case_id = case['num']
        url = case['url']
        data = case.get('data')
        data = eval(data)
        expected_result = case.get('expected_result')
        expected_result = expected_result.replace('null', 'None')
        expected_result = eval(expected_result)
        expected_msg = expected_result['msg']
        real_result = web_get(data, url)
        real_msg = real_result['msg']
        if real_msg == expected_msg:
            print('第{}条正常'.format(case_id))
            final_result = '正确'
            success += 1
        else:
            print('第{}条错误'.format(case_id))
            final_result = '错误'
            failed += 1
        write_result(filename, formname, case_id + 1, 8,
                     final_result)
    print('用例一共{}条，成功{}条，失败{}条，成功写入文件'.format(success + failed, success, failed))


if __name__ == '__main__':
    exec_cases(r'D:\programming\programming_workingspace\exercise\Python\test2.xlsx', 'register')